#_*_coding_:utf-8_*_
# Author：ibell
# Create_time: 2020-01-04 10:33 
# File: do_excel.py


from openpyxl import  load_workbook
from API_AUTO.tools.read_config import ReadConfig
from API_AUTO.tools import project_path


class DoExcel:
    """处理Excel中的测试数据"""

    def get_data(self,file_name):
        wb=load_workbook(file_name)
        mode=eval(ReadConfig.get_config(project_path.config_path,'MODE','mode'))
        test_data = []

        for key in mode:#遍历这个存在配置文件里面的字典
            sheet = wb[key]#表单名
            if mode[key]=='all':
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data['code']=sheet.cell(i,1).value
                    row_data['module'] = sheet.cell(i, 2).value
                    row_data['title'] = sheet.cell(i, 3).value
                    row_data['url'] = sheet.cell(i, 4).value
                    row_data['data'] = sheet.cell(i, 5).value
                    row_data['method'] = sheet.cell(i, 6).value
                    row_data['expected'] = sheet.cell(i, 7).value
                    row_data['check_database']=sheet.cell(i,8).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)


            else:
                for case_id in mode[key]:#[1,3]
                    row_data = {}
                    row_data['code'] = sheet.cell(case_id+1, 1).value
                    row_data['module'] = sheet.cell(case_id+1, 2).value
                    row_data['title'] = sheet.cell(case_id+1, 3).value
                    row_data['url'] = sheet.cell(case_id+1, 4).value
                    row_data['data'] = sheet.cell(case_id+1, 5).value
                    row_data['method'] = sheet.cell(case_id+1, 6).value
                    row_data['expected'] = sheet.cell(case_id+1, 7).value
                    row_data['check_database'] = sheet.cell(case_id, 8).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)

        for item in test_data:
            tel = wb['phone_data'].cell(2, 1).value
            # print(tel)
            if str(item).find("${tel_1}") != -1:
                item['data']=eval(str(item['data']).replace("${tel_1}",str(tel)))
                tel=tel+1
            elif str(item).find("${tel}") != -1:
                item['data'] = eval(str(item['data']).replace("${tel}", str(tel+1)))
                tel=tel+1
        self.updata_tel(tel, file_name, 'phone_data')
        return test_data


    def wirte_back(self,file_name,sheet_name,i,value,test_result,check_result):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,9).value=value
        sheet.cell(i, 10).value = test_result
        sheet.cell(i, 11).value = check_result
        wb.save(file_name)



    def updata_tel(self,tel,file_name,sheet_name):
        """更新Excel中的手机号"""
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(file_name)


if __name__=="__main__":
    test_data=DoExcel().get_data(project_path.test_case_path)
    print(test_data)
    # print(type(eval(test_data[1]['data'])['amount']))
