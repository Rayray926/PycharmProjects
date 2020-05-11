#_*_coding_:utf-8_*_
# Author：ibell
# Create_time: 2020-04-01 14:36 
# File: do_excel.py
import os
from openpyxl import load_workbook
from conf import config
from common.my_log import Mylog





class DoExcel:
    """
    封装所有和excel交互的操作
    """

    def __init__(self,file_path):
        """
        初始化函数，判断测试用例文件是否存在，存在进行初始化，否则抛出异常
        :param file_path: 测试用例文件路径
        """
        if os.path.exists(file_path):
            self.file_path=file_path
            self.wb=load_workbook(self.file_path)
            self.module=config.ConfigYaml().get_test_module()
            Mylog().info("需要运行的模块是{},类型是{}".format(self.module,type(self.module)))
            self.test_data = []
        else:
            Mylog().error("{}文件不存在".format(file_path))
            raise FileNotFoundError("文件不存在")


    def get_isRun_data(self):
        """
        获取测试用例数据
        :return: 根据用户配置，读取对应模块的测试用例，根据用例中的是否执行判断是否将用例加入到self.test_data 中
        """
        try:
            for sheet_name in self.module:
                sheet=self.wb[sheet_name]
                for i in range(2,sheet.max_row):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['module'] = sheet.cell(i, 2).value
                    row_data['title'] = sheet.cell(i, 3).value
                    row_data['isrun'] = sheet.cell(i, 4).value
                    row_data['url'] = sheet.cell(i, 5).value
                    row_data['pre_conditon'] = sheet.cell(i, 6).value
                    row_data['remark'] = sheet.cell(i, 7).value
                    row_data['headers'] = sheet.cell(i, 8).value
                    row_data['cookies'] = sheet.cell(i, 9).value
                    row_data['method'] = sheet.cell(i, 10).value
                    row_data['param_type'] = sheet.cell(i, 11).value
                    row_data['data'] = sheet.cell(i, 12).value
                    row_data['check_database'] = sheet.cell(i, 13).value
                    row_data['expected'] = sheet.cell(i, 14).value
                    if row_data['isrun'] == 'y':
                        self.test_data.append(row_data)
            Mylog().info("获取到的需要执行的测试用例是{}".format(self.test_data))

        except Exception as e :
            Mylog().error(e)
            raise  e

        self.test_data=self.replace_placeholder("register_tel","common_data",1,2)
        return self.test_data

    def replace_placeholder(self,type,sheet_name,row_no,col_no):
        """
        判断测试数据data中是否有手机号占位符，有则替换并更新common_data中手机号数据
        :param sheet_name:sheet名称
        :param row_no: 行号
        :param col_no:列号
        :return:
        """
        try:
            if type=="register_tel":
                register_tel=self.wb[sheet_name].cell(1,2).value
                Mylog().info("获取到的register_tel是{}".format(register_tel))
                for item in self.test_data:
                    if str(item).find('${register_tel}') != -1:
                        item['data'] = str(item['data']).replace('${register_tel}', str(register_tel))
                    elif str(item).find('${register_tel_new}') != -1:
                        item['data'] = str(item['data']).replace('${tel_1}', str(register_tel + 1))
                self.wb[sheet_name].cell(row_no, col_no).value=register_tel+1

            self.wb.save(self.file_path)
        except Exception as e:
            Mylog().error("replace_placeholder {}".format(e))
            raise e

        return self.test_data


    def get_module_test_data(self,module_name):
        self.test_data=self.get_isRun_data()
        self.test_module_data=[]
        for item in self.test_data:
            if item["module"]==module_name:
                self.test_module_data.append(item)
        Mylog().info("获取的测试模块是{}，数据是{}".format(module_name,self.test_module_data))
        return  self.test_module_data


    def write_back(self,sheet_name,row_no,res_data,check_result,test_result):
        sheet = self.wb[sheet_name]
        sheet.cell(row_no, 15).value = res_data
        sheet.cell(row_no, 16).value = check_result
        sheet.cell(row_no, 17).value = test_result
        self.wb.save(self.file_path)


if __name__ == '__main__':
    # t=DoExcel(config.get_test_data_path()).get_module_test_data("withdraw")
    test_data_path = config.get_test_data_path()
    test_data = DoExcel(test_data_path).get_module_test_data('withdraw')
    print(test_data)





